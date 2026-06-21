"""SEC EDGAR 10-Q/10-K Excel 确定性解析 → extracted_json。"""
import io
import re
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

from app.services.financial_ai import normalize_extracted_payload
from app.services.fiscal_calendar import build_period_context, _parse_date

_INCOME_MAP = {
    "revenue": ("revenue",),
    "cost of goods sold": ("cogs",),
    "gross margin": ("gross_profit",),
    "research and development": ("rd",),
    "selling, general, and administrative": ("sga",),
    "operating income": ("operating_income",),
    "operating income (loss)": ("operating_income",),
    "income tax": ("tax",),
    "income tax (provision) benefit": ("tax",),
    "net income": ("net_income",),
    "net income (loss)": ("net_income",),
}

_BALANCE_MAP = {
    "cash and cash equivalents": ("cash",),
    "receivables": ("receivables",),
    "inventories": ("inventory",),
    "property, plant, and equipment": ("ppe",),
    "total assets": ("total_assets",),
    "total current liabilities": ("current_liabilities",),
    "long-term debt": ("long_term_debt",),
    "total liabilities": ("total_liabilities",),
    "total shareholders' equity": ("equity",),
    "total shareholders’ equity": ("equity",),
    "total stockholders' equity": ("equity",),
    "total stockholders’ equity": ("equity",),
}

_CASH_MAP = {
    "net cash provided by operating activities": ("operating",),
    "net cash used for investing activities": ("investing",),
    "net cash provided by (used for) financing activities": ("financing",),
}


def _normalize_label(text: Any) -> str:
    s = str(text or "").strip().lower()
    s = s.replace("\u2019", "'").replace("\u2018", "'")
    s = re.sub(r"\s+", " ", s)
    return s


def _to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if value != value:  # NaN
            return None
        return float(value)
    text = str(value).strip().replace(",", "").replace("—", "").replace("–", "")
    if not text or text in ("-", "—"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _cik_from_filename(filename: str) -> Optional[str]:
    match = re.search(r"(\d{10})", filename or "")
    return match.group(1) if match else None


def _detect_form_type(sheet_names: List[str]) -> str:
    joined = " ".join(sheet_names).lower()
    if re.search(r"\b10\s*-?\s*k\b", joined):
        return "10-K"
    if re.search(r"\b10\s*-?\s*q\b", joined):
        return "10-Q"
    return "10-Q"


def _find_sheet_name(sheet_names: List[str], *keywords: str) -> Optional[str]:
    lower_map = {name.lower(): name for name in sheet_names}
    for key in keywords:
        key_l = key.lower()
        for low, orig in lower_map.items():
            if key_l in low:
                return orig
    return None


def _find_cash_flow_sheet(workbook) -> Optional[str]:
    """优先完整 Consolidated Statements of Cash Flows 表。"""
    candidates = []
    for name in workbook.sheet_names():
        low = name.lower()
        if "cash flow" not in low:
            continue
        sh = workbook.sheet_by_name(name)
        header = " ".join(
            _normalize_label(sh.cell_value(r, c))
            for r in range(min(5, sh.nrows))
            for c in range(min(sh.ncols, 8))
        )
        score = 0
        if "consolidated statements of cash flows" in header:
            score += 10
        if "nine months" in header or "three months" in header or "quarter ended" in header:
            score += 3
        if "for the year ended" in header:
            score += 2
        candidates.append((score, name))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def _load_workbook(data: bytes, filename: str):
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    if ext == "xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

        class _Cell:
            def __init__(self, value):
                self.value = value

        class _Sheet:
            def __init__(self, ws):
                self._ws = ws
                self.nrows = ws.max_row or 0
                self.ncols = ws.max_column or 0

            def cell_value(self, row, col):
                if row >= self.nrows or col >= self.ncols:
                    return ""
                val = self._ws.cell(row=row + 1, column=col + 1).value
                return "" if val is None else val

        class _Workbook:
            def __init__(self, xlsx_wb):
                self._sheets = xlsx_wb.sheetnames
                self._wb = xlsx_wb

            def sheet_names(self):
                return list(self._sheets)

            def sheet_by_name(self, name):
                return _Sheet(self._wb[name])

            def sheet_by_index(self, idx):
                return _Sheet(self._wb[self._sheets[idx]])

        return _Workbook(wb), "xlsx"

    import xlrd

    return xlrd.open_workbook(file_contents=data), "xls"


def _cell_value(sh, row: int, col: int) -> Any:
    try:
        val = sh.cell_value(row, col)
    except Exception:
        return ""
    if val is None:
        return ""
    try:
        import xlrd

        if hasattr(sh, "cell_type") and sh.cell_type(row, col) == xlrd.XL_CELL_DATE:
            from xlrd.xldate import xldate_as_datetime

            return xldate_as_datetime(val, False).date().isoformat()
    except Exception:
        pass
    return val


def _sheet_text_matrix(sh, max_rows: int = 120, max_cols: int = 16) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for r in range(min(max_rows, sh.nrows)):
        row = [_cell_value(sh, r, c) for c in range(min(max_cols, sh.ncols))]
        rows.append(row)
    return rows


def _extract_dates_from_header(rows: List[List[Any]]) -> Tuple[List[Optional[date]], str]:
    """返回每列对应日期、表头 scope 描述。"""
    scope = "unknown"
    header_blob = " ".join(_normalize_label(v) for row in rows[:8] for v in row)
    if "nine months ended" in header_blob:
        scope = "ytd"
    elif "three months ended" in header_blob or "quarter ended" in header_blob:
        scope = "quarter"
    elif "for the year ended" in header_blob or "years ended" in header_blob:
        scope = "annual"

    dates: List[Optional[date]] = []
    for row in rows[:8]:
        for c, cell in enumerate(row):
            while len(dates) <= c:
                dates.append(None)
            if dates[c] is not None:
                continue
            parsed = _parse_date(cell)
            if parsed:
                dates[c] = parsed
            elif isinstance(cell, str) and re.search(r"\b20\d{2}\b", cell):
                combined = cell
                if c + 1 < len(row):
                    combined = f"{cell} {row[c + 1]}"
                parsed = _parse_date(combined)
                if parsed:
                    dates[c] = parsed
    return dates, scope


def _pick_columns(dates: List[Optional[date]]) -> Tuple[int, Optional[int]]:
    """当前列、同比列索引。"""
    valid = [(i, d) for i, d in enumerate(dates) if d is not None]
    if not valid:
        return 1, None
    valid.sort(key=lambda x: x[1], reverse=True)
    current_col = valid[0][0]
    prior_col = valid[1][0] if len(valid) > 1 else None
    return current_col, prior_col


def _match_field(label: str, mapping: Dict[str, Tuple[str, ...]]) -> Optional[str]:
    norm = _normalize_label(label)
    if not norm:
        return None
    for key, fields in mapping.items():
        if key in norm or norm.startswith(key):
            return fields[0]
    return None


def _parse_statement_rows(
    rows: List[List[Any]],
    mapping: Dict[str, Tuple[str, ...]],
) -> Tuple[Dict[str, float], Dict[str, float], str, Optional[date]]:
    dates, scope = _extract_dates_from_header(rows)
    current_col, prior_col = _pick_columns(dates)
    current: Dict[str, float] = {}
    prior: Dict[str, float] = {}

    for row in rows:
        if not row:
            continue
        label = row[0] if row else ""
        field = _match_field(str(label), mapping)
        if not field:
            continue
        cur_val = _to_float(row[current_col] if current_col < len(row) else None)
        if cur_val is not None:
            current[field] = round(cur_val, 2)
        if prior_col is not None and prior_col < len(row):
            prev_val = _to_float(row[prior_col])
            if prev_val is not None:
                prior[field] = round(prev_val, 2)

    period_end = None
    valid_dates = [d for d in dates if d]
    if valid_dates:
        period_end = max(valid_dates)
    return current, prior, scope, period_end


def _yoy_pct(current: Optional[float], prior: Optional[float]) -> Optional[float]:
    if current is None or prior is None or prior == 0:
        return None
    return round((current - prior) / abs(prior) * 100, 2)


def _build_kpis(
    income: Dict[str, float],
    income_prior: Dict[str, float],
    cash_flow: Dict[str, float],
    cash_scope: str,
) -> Dict[str, Any]:
    kpis: Dict[str, Any] = {}
    rev = income.get("revenue")
    if rev is not None:
        kpis["revenue"] = {
            "value": rev,
            "yoy_pct": _yoy_pct(rev, income_prior.get("revenue")),
            "qoq_pct": None,
        }
    net = income.get("net_income")
    if net is not None:
        kpis["net_profit"] = {
            "value": net,
            "yoy_pct": _yoy_pct(net, income_prior.get("net_income")),
            "qoq_pct": None,
        }
    if income.get("revenue") and income.get("gross_profit") is not None:
        kpis["gross_margin_pct"] = round(
            income["gross_profit"] / income["revenue"] * 100, 2
        )
    if cash_scope == "quarter" and cash_flow.get("operating") is not None:
        kpis["operating_cf"] = cash_flow["operating"]
        kpis["free_cash_flow"] = cash_flow["operating"]
    return kpis


def parse_sec_filing(
    data: bytes,
    filename: str = "",
    *,
    ticker: str | None = None,
) -> Dict[str, Any]:
    """解析 SEC xls/xlsx，返回 extracted、filing_meta、parse_log、source_text_summary。"""
    parse_log: List[str] = []
    wb, fmt = _load_workbook(data, filename)
    sheet_names = wb.sheet_names()
    parse_log.append(f"format={fmt}, sheets={len(sheet_names)}")

    form_type = _detect_form_type(sheet_names)
    cik = _cik_from_filename(filename)
    company_sheet = _find_sheet_name(sheet_names, "technology", "inc", "corp", "company") or sheet_names[0]
    company_name = ""
    try:
        sh0 = wb.sheet_by_name(company_sheet)
        for r in range(min(5, sh0.nrows)):
            val = str(sh0.cell_value(r, 0) or sh0.cell_value(r, 1) or "").strip()
            if val and len(val) > 3 and not val.lower().startswith("delaware"):
                company_name = val
                break
    except Exception:
        pass

    ops_name = _find_sheet_name(sheet_names, "operations")
    bal_name = _find_sheet_name(sheet_names, "consolidated balance sheets", "balance sheets")
    cf_name = _find_cash_flow_sheet(wb) if hasattr(wb, "sheet_names") else _find_sheet_name(sheet_names, "cash flows")

    income, income_prior, income_scope, period_end = {}, {}, "unknown", None
    balance, balance_prior, balance_scope, balance_end = {}, {}, "unknown", None
    cash_flow, cash_prior, cash_scope, cash_end = {}, {}, "unknown", None

    if ops_name:
        rows = _sheet_text_matrix(wb.sheet_by_name(ops_name))
        income, income_prior, income_scope, period_end = _parse_statement_rows(rows, _INCOME_MAP)
        parse_log.append(f"operations: {len(income)} fields, scope={income_scope}")
    else:
        parse_log.append("missing operations sheet")

    if bal_name:
        rows = _sheet_text_matrix(wb.sheet_by_name(bal_name))
        balance, balance_prior, balance_scope, balance_end = _parse_statement_rows(rows, _BALANCE_MAP)
        parse_log.append(f"balance: {len(balance)} fields")
        if balance.get("equity") is None and balance.get("total_assets") and balance.get("total_liabilities"):
            balance["equity"] = round(balance["total_assets"] - balance["total_liabilities"], 2)
            parse_log.append("equity derived from assets - liabilities")
        balance.pop("total_liabilities", None)
    else:
        parse_log.append("missing balance sheet")

    if cf_name:
        rows = _sheet_text_matrix(wb.sheet_by_name(cf_name))
        cash_flow, cash_prior, cash_scope, _ = _parse_statement_rows(rows, _CASH_MAP)
        parse_log.append(f"cash flows: {len(cash_flow)} fields, scope={cash_scope}")
    else:
        parse_log.append("missing cash flows sheet")

    end_date = period_end or balance_end or cash_end
    if not end_date:
        raise ValueError("无法从财报中识别报告期末日期")

    period_ctx = build_period_context(end_date, ticker=ticker)
    calendar_period = period_ctx["calendar_period"]

    use_cash = cash_scope in ("quarter", "annual") and cash_flow
    cash_flow_block: Dict[str, Dict[str, float]] = {}
    if use_cash:
        cash_flow_block[calendar_period] = cash_flow

    raw_payload: Dict[str, Any] = {
        "currency": "USD",
        "unit": "millions",
        "periods": [calendar_period],
        "filing_meta": {
            "source": "sec_xls",
            "form_type": form_type,
            "period_end": period_ctx["period_end"],
            "calendar_period": calendar_period,
            "filing_fy": period_ctx["filing_fy"],
            "filing_fq": period_ctx["filing_fq"],
            "fiscal_year_end_month": period_ctx["fiscal_year_end_month"],
            "cash_flow_scope": cash_scope,
            "income_scope": income_scope,
            "cik": cik,
            "company_name": company_name or None,
        },
        "income_statement": {calendar_period: income} if income else {},
        "balance_sheet": {calendar_period: balance} if balance else {},
        "cash_flow": cash_flow_block,
        "kpis": {
            calendar_period: _build_kpis(income, income_prior, cash_flow if use_cash else {}, cash_scope),
        },
        "red_flags": [],
        "material_events": [],
        "ai_summary": "",
    }

    historical: List[Dict[str, Any]] = []
    if income_prior:
        historical.append({"scope": "prior_year_column", "income_statement": income_prior})
    if balance_prior:
        if historical:
            historical[0]["balance_sheet"] = balance_prior
        else:
            historical.append({"scope": "prior_year_column", "balance_sheet": balance_prior})
    if historical:
        raw_payload["historical_periods"] = historical

    extracted = normalize_extracted_payload(raw_payload)
    extracted["filing_meta"] = raw_payload["filing_meta"]
    if historical:
        extracted["historical_periods"] = historical

    summary_lines = [
        f"{form_type} · {company_name or ticker or 'SEC'}",
        f"Period end: {period_ctx['period_end']} → {calendar_period}",
        f"FY{period_ctx['filing_fy']} Q{period_ctx['filing_fq']}",
        f"Revenue: {income.get('revenue')} | Net income: {income.get('net_income')}",
        f"Cash flow scope: {cash_scope}",
    ]

    return {
        "extracted": extracted,
        "filing_meta": raw_payload["filing_meta"],
        "parse_log": parse_log,
        "source_text_summary": "\n".join(summary_lines),
        "suggested_ticker": (ticker or "").upper() or None,
        "suggested_fiscal_period": calendar_period,
        "suggested_report_date": period_ctx["period_end"],
        "suggested_title": f"{(ticker or company_name or 'SEC').strip()} {calendar_period} {form_type}",
    }
